import time
import openpyxl
from csv import writer
from hashlib import new

# import results as results
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.wait import WebDriverWait

driver = webdriver.Chrome(
        executable_path="C:\\Users\\neetesh\\Downloads\\chromedriver_win32 (2)\\chromedriver.exe")
print(type(driver))
driver.get("https://www.jobtrees.com/")
driver.maximize_window()

wb = load_workbook('C:\\Users\\neetesh\\Downloads\\Test-Automation\\Avgtimerole.xlsx')

ws = wb.active


# print(f'{ws["A2"].value}:{ws["2"].value}')
column_a = ws['A']

# print(column_a)
# For loop
filename = "C:\\Users\\neetesh\\Downloads\\Test-Automation\\Avgtimerole.xlsx"
workbook = openpyxl.load_workbook(filename)
worksheet = workbook.active

# row = 1

for row in range(2, len(column_a)+1):
    driver.find_element(By.CLASS_NAME, "search-nav").click()
    driver.find_element(By.ID, "search-suggestions-input").send_keys(worksheet.cell(column=1, row=row).value)
    driver.find_element(By.CLASS_NAME, "search-input").send_keys(Keys.ENTER)
    time.sleep(5)

# write excel
    if driver.find_element(By.CLASS_NAME, "avg-time-in-role").text == "0":
    # if "0" in driver.find_element(By.CLASS_NAME, "avg-time-in-role").text:
        worksheet.cell(column=2, row=row).value = "Fail"
        row = row + 1
    else:
        worksheet.cell(column=2, row=row).value = "Pass"
        row = row + 1
    workbook.save(filename)
